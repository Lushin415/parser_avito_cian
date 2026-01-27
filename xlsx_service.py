import os
from openpyxl import Workbook, load_workbook
from datetime import datetime
from tzlocal import get_localzone
from models import Item
from cian_models import CianItem
from typing import Union


class XLSXHandler:
    """Сохраняет информацию в xlsx (поддерживает Avito и Cian)"""

    def __init__(self, file_name):
        self._initialize(file_name=file_name)

    def _initialize(self, file_name):
        self.file_name = file_name
        os.makedirs("result", exist_ok=True)
        if not os.path.exists(self.file_name):
            self._create_file()

    def _create_file(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Data"

        # Универсальные колонки для обоих источников
        sheet.append([
            "Источник",  # Avito или Cian
            "ID",
            "Название",
            "Цена (руб/мес)",
            "Площадь (м²)",  # Для Cian
            "URL",
            "Описание",
            "Дата публикации",  # Для Avito
            "Продавец/Автор",
            "Тип автора",  # owner/agent/agency
            "Адрес",
            "Округ",  # Для Cian (ЦАО, ЗАО и т.д.)
            "Район",
            "Метро",
            "Удалённость от метро",  # Для Cian
            "Улица",
            "Дом",
            "Координаты",  # Для Avito
            "Изображения",  # Для Avito
            "Поднято",  # Для Avito
            "Просмотры (всего)",  # Для Avito
            "Просмотры (сегодня)",  # Для Avito
        ])
        workbook.save(self.file_name)

    @staticmethod
    def get_ad_time(ad: Item):
        """Получает время публикации (только для Avito)"""
        return datetime.fromtimestamp(ad.sortTimeStamp / 1000, tz=get_localzone()).replace(tzinfo=None)

    @staticmethod
    def get_item_coords(ad: Item) -> str:
        """Возвращает координаты (только для Avito)"""
        if ad.coords and 'lat' in ad.coords and 'lng' in ad.coords:
            return f"{ad.coords['lat']};{ad.coords['lng']}"
        return ""

    @staticmethod
    def get_item_address_user(ad: Item) -> str:
        """Возвращает адрес пользователя (только для Avito)"""
        if ad.coords and 'address_user' in ad.coords:
            return ad.coords['address_user']
        return ""

    def append_data_from_page(self, ads: list[Union[Item, CianItem]]):
        """Добавляет данные из списка объявлений (Avito или Cian)"""
        workbook = load_workbook(self.file_name)
        sheet = workbook.active

        for ad in ads:
            if isinstance(ad, Item):  # Avito
                row = self._format_avito_row(ad)
            elif isinstance(ad, CianItem):  # Cian
                row = self._format_cian_row(ad)
            else:
                continue  # Неизвестный тип - пропускаем

            sheet.append(row)

        workbook.save(self.file_name)

    def _format_avito_row(self, ad: Item) -> list:
        """Форматирует строку для Avito"""

        # Получаем изображения
        def get_largest_image_url(img):
            best_key = max(
                img.root.keys(),
                key=lambda k: int(k.split("x")[0]) * int(k.split("x")[1])
            )
            return str(img.root[best_key])

        images_urls = [get_largest_image_url(img) for img in ad.images] if ad.images else []

        return [
            "Avito",  # Источник
            ad.id,  # ID
            ad.title,  # Название
            ad.priceDetailed.value if ad.priceDetailed else "",  # Цена
            "",  # Площадь (нет у Avito)
            f"https://www.avito.ru/{ad.urlPath}",  # URL
            ad.description if ad.description else "",  # Описание
            self.get_ad_time(ad=ad),  # Дата публикации
            ad.sellerId if ad.sellerId else "",  # Продавец
            "",  # Тип автора (нет у Avito в таком виде)
            ad.location.name if ad.location else "",  # Адрес (город)
            "",  # Округ (нет у Avito)
            "",  # Район (нет структурировано)
            "",  # Метро (нет структурировано)
            "",  # Удалённость (нет)
            "",  # Улица (нет структурировано)
            "",  # Дом (нет структурировано)
            self.get_item_coords(ad=ad),  # Координаты
            ";".join(images_urls),  # Изображения
            "Да" if ad.isPromotion else "Нет",  # Поднято
            ad.total_views if ad.total_views is not None else "",  # Просмотры всего
            ad.today_views if ad.today_views is not None else "",  # Просмотры сегодня
        ]

    def _format_cian_row(self, ad: CianItem) -> list:
        """Форматирует строку для Cian"""
        return [
            "Cian",  # Источник
            ad.id,  # ID
            ad.title,  # Название
            ad.price.value if ad.price else 0,  # Цена
            ad.total_meters if ad.total_meters > 0 else "",  # Площадь
            ad.url,  # URL
            ad.description[:500] if ad.description else "",  # Описание (обрезаем)
            "",  # Дата публикации (нет у Cian в списке)
            ad.author.name if ad.author else "",  # Автор
            ad.author.type if ad.author else "",  # Тип автора
            ad.location,  # Адрес (город)
            ad.location_data.district_okrug if ad.location_data else "",  # Округ
            ad.location_data.district if ad.location_data else "",  # Район
            ad.location_data.underground if ad.location_data else "",  # Метро
            ad.location_data.metro_remoteness if ad.location_data else "",  # Удалённость
            ad.location_data.street if ad.location_data else "",  # Улица
            ad.location_data.house_number if ad.location_data else "",  # Дом
            "",  # Координаты (нет в списке)
            "",  # Изображения (нет в списке)
            "",  # Поднято (нет у Cian)
            "",  # Просмотры всего (нет в списке)
            "",  # Просмотры сегодня (нет в списке)
        ]